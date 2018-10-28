import openpyxl
import random

def allocate_full(filename):
    dsa_allocation(filename)
    mep_allocation(filename)
    allocation_1stchoice_ranks(filename)
    remove_extra_full(filename)
    allocation_2ndchoice_ranks(filename)
    remove_extra_full(filename)
    allocation_3rdchoice_ranks(filename)
    remove_extra_full(filename)
    allocate_remaining_choices(filename)
    allocate_remaining(filename)

def dsa_allocation(filename): #Allocation for DSA
    wb = openpyxl.load_workbook(filename)
    DSA_sheet = wb["DSA"]
    class_sheet = wb["classlist"]
    dsa_dict = {}
    for i in range(1,DSA_sheet.max_row):
        dsa_dict[DSA_sheet.cell(row = i+1, column = 1).value] = DSA_sheet.cell(row = i+1, column = 3).value
    for key in dsa_dict:
        for i in range(1,class_sheet.max_row):
            if key == class_sheet.cell(row = i, column = 4).value:
                class_sheet["G" + str(i)] = dsa_dict[key]
                class_sheet["H" + str(i)] = "DSA"
                class_sheet["I" + str(i)] = ""
            else:
                continue
    wb.save(filename)

def mep_allocation(filename): #Allocation for MEP to their 1st choice Music CCA
    wb = openpyxl.load_workbook(filename)
    mep_sheet = wb["MEP"]
    choice_sheet = wb["choices"]
    class_sheet = wb["classlist"]
    mep_namelst = []
    mep_choice_dict = {}
    for i in range(1,mep_sheet.max_row):
        mep_namelst.append(mep_sheet.cell(row = i+1, column = 1).value)
    for i in mep_namelst:
        for j in range(1,choice_sheet.max_row):
            if i == choice_sheet.cell(row = j+1, column = 2).value:
                mep_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 6).value,choice_sheet.cell(row = j+1, column = 7).value]
    for key,value in mep_choice_dict.items():
        for i in range(1,class_sheet.max_row):
            if key == class_sheet.cell(row = i+1, column = 4).value:
                if value[0] == "RV" or value[0] == "SE" or value[0] == "RICO" or value[0] == "RIMB" or value[0] == "GE":
                    class_sheet["G" + str(i+1)] = value[0]
                    class_sheet["H" + str(i+1)] = 1
                    class_sheet["I" + str(i+1)] = ""
                elif value[1] == "RV" or value[1] == "SE" or value[1] == "RICO" or value[1] == "RIMB" or value[1] == "GE":
                    class_sheet["G" + str(i+1)] = value[1]
                    class_sheet["H" + str(i+1)] = 2
                    class_sheet["I" + str(i+1)] = ""
            else:
                continue


def allocation_1stchoice_ranks(filename): #Allocate students to their 1st choice CCA if they have rankings
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    cca_ranking_dict = cca_ranking_compilation(filename)
    remaining_choice_dict = remaining_people(filename,1,"")
    choice_sheet = wb["choices"]
    for key, value in remaining_choice_dict.items():
        for cca, ranking in cca_ranking_dict.items():
            if value[0] == cca:
                for rank, name in ranking.items():
                    if name[0] == key and name[1] == value[1]:
                        for i in range(1,class_sheet.max_row):
                            if name[0] == class_sheet.cell(row = i+1, column = 4).value:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 1
                                class_sheet["I" + str(i+1)] = rank
                            else:
                                continue
                    else:
                        continue
            else:
                continue
    wb.save(filename)

def remove_extra_full(filename): #Remove extra students from all CCAs that go beyond quota
    cca_extra = extra_cca(filename,"extra")
    if "BAD" in cca_extra:
        remove_extra(filename,"BAD")
    if "BAS" in cca_extra:
        remove_extra(filename,"BAS")
    if "CRI" in cca_extra:
        remove_extra(filename,"CRI")
    if "CC" in cca_extra:
        remove_extra(filename,"CC")
    if "FEN" in cca_extra:
        remove_extra(filename,"FEN")
    if "FLO" in cca_extra:
        remove_extra(filename,"FLO")
    if "GOLF" in cca_extra:
        remove_extra(filename,"GOLF")
    if "HOC" in cca_extra:
        remove_extra(filename,"HOC")
    if "JUD" in cca_extra:
        remove_extra(filename,"RUG")
    if "SHO" in cca_extra:
        remove_extra(filename,"SHO")
    if "SQU" in cca_extra:
        remove_extra(filename,"SQU")
    if "SOF" in cca_extra:
        remove_extra(filename,"SOF")
    if "TAB" in cca_extra:
        remove_extra(filename,"TAB")
    if "TEN" in cca_extra:
        remove_extra(filename,"TEN")
    if "TNF" in cca_extra:
        remove_extra(filename,"TNF")
    if "POLO" in cca_extra:
        remove_extra(filename,"POLO")
    if "01SCOUT" in cca_extra:
        remove_extra(filename,"01SCOUT")
    if "02SCOUT" in cca_extra:
        remove_extra(filename,"02SCOUT")
    if "BB" in cca_extra:
        remove_extra(filename,"BB")
    if "NCC" in cca_extra:
        remove_extra(filename,"NCC")
    if "NPCC" in cca_extra:
        remove_extra(filename,"NPCC")
    if "RC" in cca_extra:
        remove_extra(filename,"RC")
    if "RICO" in cca_extra:
        remove_extra(filename,"RICO")
    if "GE" in cca_extra:
        remove_extra(filename,"GE")
    if "RIMB" in cca_extra:
        remove_extra(filename,"RIMB")
    if "RV" in cca_extra:
        remove_extra(filename,"RV")
    if "RP" in cca_extra:
        remove_extra(filename,"RP")
    if "SE" in cca_extra:
        remove_extra(filename,"SE")
    if "Debater" in cca_extra:
        remove_extra(filename,"Debater")

def remove_extra(filename,cca): #Remove extra students from individual CCA that go beyond quota based on rank
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    ccaquota_sheet = wb["ccaquota"]
    dsa_count = dsa_counter(filename,cca)
    assigned = []
    remaining = []
    ccaquota = 0
    for i in range(1,class_sheet.max_row):
        if class_sheet.cell(row = i+1, column = 7).value == cca:
            if class_sheet.cell(row = i+1, column = 9).value != None:
                assigned.append([class_sheet.cell(row = i+1, column = 3).value,class_sheet.cell(row = i+1, column = 9).value])
    for i in range(1,ccaquota_sheet.max_row):
        if ccaquota_sheet.cell(row = i+2, column = 2).value == cca:
            ccaquota += int(ccaquota_sheet.cell(row = i+2, column = 5).value)

    for j in range((ccaquota-dsa_count),len(assigned)):
        remaining.append(sorted(assigned,key=lambda rec:int(rec[1]))[j])
    for i in remaining:
        for j in range(1,class_sheet.max_row):
            if i[0] == class_sheet.cell(row = j+1, column = 3).value:
                class_sheet["G" + str(j+1)] = ""
                class_sheet["H" + str(j+1)] = ""
                class_sheet["I" + str(j+1)] = ""
    wb.save(filename)
    return remaining

def extra_cca(filename,option): #Returns list of CCAs that exceed quota OR return dictionary of the quota and occupied positions of all CCAs
    wb = openpyxl.load_workbook(filename)
    ccaquota_sheet = wb["ccaquota"]
    class_sheet = wb["classlist"]
    CCA_extra = []
    ccaquota_dict = {}
    assigned = []
    CCA_namelist = ["BAS", "BAD", "CRI", "CC", "FEN", "FLO", "GOLF", "HOC", "JUD", "RUG", "SHO", "SOF", "SQU", "TAB", "TEN", "TNF", "POLO", "BB", "01SCOUT","02SCOUT", "NCC", "NPCC", "RC", "RICO", "GE","RIMB", "RV", "RP", "SE", "Debater"]
    for i in CCA_namelist:
        for j in range(1,class_sheet.max_row):
            if class_sheet.cell(row = j+1, column = 7).value == i:
                assigned.append(class_sheet.cell(row = j+1, column = 3).value)
        for x in range(1,ccaquota_sheet.max_row):
            if ccaquota_sheet.cell(row = x+2, column = 2).value == i:
                ccaquota_dict[i] = [ccaquota_sheet.cell(row = x+2, column = 5).value,len(assigned)]
        assigned = []
    if option == "extra":
        for key,value in ccaquota_dict.items():
            if int(value[1]) > int(value[0]):
                CCA_extra.append(key)
        return CCA_extra
    elif option == "quota":
        return ccaquota_dict

def dsa_counter(filename,cca): #Counts the number of DSA students in particular CCA
    wb = openpyxl.load_workbook(filename)
    DSA_sheet = wb["DSA"]
    counter = 0
    for i in range(1,DSA_sheet.max_row):
        if DSA_sheet.cell(row = i+1, column = 3).value == cca:
            counter +=1
    return counter

def allocation_2ndchoice_ranks(filename): #Allocate students to their 2nd choice CCA if they have rankings
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,2,"")
    remaining_name_lst = []
    cca_ranking_dict = cca_ranking_compilation(filename)
    choice_sheet = wb["choices"]

    for key, value in remaining_choice_dict.items():
        for cca, ranking in cca_ranking_dict.items():
            if value[0] == cca:
                for rank, name in ranking.items():
                    if name[0] == key and name[1] == value[1]:
                        for i in range(1,class_sheet.max_row):
                            if name[0] == class_sheet.cell(row = i+1, column = 4).value:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 2
                                class_sheet["I" + str(i+1)] = rank
                            else:
                                continue
                    else:
                        continue
            else:
                continue

    wb.save(filename)

def allocation_3rdchoice_ranks(filename): #Allocate students to their 3rd choice CCA if they have rankings
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,3,"")
    remaining_name_lst = []
    cca_ranking_dict = cca_ranking_compilation(filename)
    choice_sheet = wb["choices"]

    for key, value in remaining_choice_dict.items():
        for cca, ranking in cca_ranking_dict.items():
            if value[0] == cca:
                for rank, name in ranking.items():
                    if name[0] == key and name[1] == value[1]:
                        for i in range(1,class_sheet.max_row):
                            if name[0] == class_sheet.cell(row = i+1, column = 4).value:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 3
                                class_sheet["I" + str(i+1)] = rank
                            else:
                                continue
                    else:
                        continue
            else:
                continue

    wb.save(filename)

def cca_ranking_compilation(filename): #Compilation of rankings in all CCAs in a dictionary
    wb = openpyxl.load_workbook(filename)
    cca_ranking_dict = {}
    cca_ranking_dict["BAS"] = cca_ranking_indiv(filename,"BAS")
    cca_ranking_dict["BAD"] = cca_ranking_indiv(filename,"BAD")
    cca_ranking_dict["CRI"] = cca_ranking_indiv(filename,"CRI")
    cca_ranking_dict["CC"] = cca_ranking_indiv(filename,"CC")
    cca_ranking_dict["FEN"] = cca_ranking_indiv(filename,"FEN")
    cca_ranking_dict["FLO"] = cca_ranking_indiv(filename,"FLO")
    cca_ranking_dict["GOLF"] = cca_ranking_indiv(filename,"GOLF")
    cca_ranking_dict["HOC"] = cca_ranking_indiv(filename,"HOC")
    cca_ranking_dict["JUD"] = cca_ranking_indiv(filename,"JUD")
    cca_ranking_dict["RUG"] = cca_ranking_indiv(filename,"RUG")
    cca_ranking_dict["SHO"] = cca_ranking_indiv(filename,"SHO")
    cca_ranking_dict["SOF"] = cca_ranking_indiv(filename,"SOF")
    cca_ranking_dict["SQU"] = cca_ranking_indiv(filename,"SQU")
    cca_ranking_dict["TAB"] = cca_ranking_indiv(filename,"TAB")
    cca_ranking_dict["TEN"] = cca_ranking_indiv(filename,"TEN")
    cca_ranking_dict["TNF"] = cca_ranking_indiv(filename,"TNF")
    cca_ranking_dict["POLO"] = cca_ranking_indiv(filename,"POLO")
    cca_ranking_dict["BB"] = cca_ranking_indiv(filename,"BB")
    cca_ranking_dict["01SCOUT"] = cca_ranking_indiv(filename,"O1")
    cca_ranking_dict["02SCOUT"] = cca_ranking_indiv(filename,"O2")
    cca_ranking_dict["NCC"] = cca_ranking_indiv(filename,"NCC")
    cca_ranking_dict["NPCC"] = cca_ranking_indiv(filename,"NPCC")
    cca_ranking_dict["RC"] = cca_ranking_indiv(filename,"RC")
    cca_ranking_dict["RICO"] = cca_ranking_indiv(filename,"CO")
    cca_ranking_dict["RIMB"] = cca_ranking_indiv(filename,"RIMB")
    cca_ranking_dict["GE"] = cca_ranking_indiv(filename,"GE")
    cca_ranking_dict["RV"] = cca_ranking_indiv(filename,"RV")
    cca_ranking_dict["RP"] = cca_ranking_indiv(filename,"RP")
    cca_ranking_dict["SE"] = cca_ranking_indiv(filename,"SE")
    cca_ranking_dict["Debater"] = cca_ranking_indiv(filename,"Debater")
    return cca_ranking_dict

def cca_ranking_indiv(filename,sheetname): #Compilation of rankings in each CCA
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    cca_ranking_dict = {}
    for i in range(1,sheet.max_row):
        cca_ranking_dict[str(sheet.cell(row = i+1, column = 1).value)] = [sheet.cell(row = i+1, column = 2).value,sheet.cell(row = i+1, column = 3).value]
    return cca_ranking_dict

def remaining_people(filename,choice,option): #Returns list of remaining unallocated students OR a dictionary of unallocated students, their choice of CCA & their class
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_name_lst = []
    remaining_choice_dict = {}
    choice_sheet = wb["choices"]
    for x in range(1,(class_sheet.max_row)): #remaining list of people who dont have ccas
        if class_sheet.cell(row = x+1, column = 7).value == None:
            remaining_name_lst.append(class_sheet.cell(row = x+1, column = 3).value)
        else:
            continue
    if option == "list":
        return remaining_name_lst
    else:
        for i in remaining_name_lst:
            for j in range(1,choice_sheet.max_row):
                if i == choice_sheet.cell(row = j+1, column = 2).value:
                    if choice == 1:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 6).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 2:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 7).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 3:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 8).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 4:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 9).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 5:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 10).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 6:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 11).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 7:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 12).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 8:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 13).value,choice_sheet.cell(row = j+1, column = 4).value]
                    elif choice == 9:
                        remaining_choice_dict[i] = [choice_sheet.cell(row = j+1, column = 14).value,choice_sheet.cell(row = j+1, column = 4).value]
        return remaining_choice_dict

def remaining_ccalist(dictionary): #Returns a list of CCAs that still have vacancies
    cca_list = []
    for key, value in dictionary.items():
        if value[0] != value[1]:
            cca_list.append(key)
        else:
            continue
    return cca_list

def allocate_remaining_choices(filename): #Allocate remaining students to their CCAs based on choice if CCAs still have vacancies
    allocate_remaining_1stchoice(filename)
    allocate_remaining_2ndchoice(filename)
    allocate_remaining_3rdchoice(filename)
    allocate_remaining_4thchoice(filename)
    allocate_remaining_5thchoice(filename)
    allocate_remaining_6thchoice(filename)
    allocate_remaining_7thchoice(filename)
    allocate_remaining_8thchoice(filename)
    allocate_remaining_9thchoice(filename)

def allocate_remaining_1stchoice(filename): #Allocate remaining students to their 1st choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,1,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 1
                                occupied += 1
    wb.save(filename)

def allocate_remaining_2ndchoice(filename): #Allocate remaining students to their 2nd choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,2,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 2
                                occupied += 1
    wb.save(filename)

def allocate_remaining_3rdchoice(filename): #Allocate remaining students to their 3rd choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,3,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 3
                                occupied += 1
    wb.save(filename)

def allocate_remaining_4thchoice(filename): #Allocate remaining students to their 4th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,4,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 4
                                occupied += 1
    wb.save(filename)

def allocate_remaining_5thchoice(filename): #Allocate remaining students to their 5th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,5,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota!= occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 5
                                occupied += 1
    wb.save(filename)

def allocate_remaining_6thchoice(filename): #Allocate remaining students to their 6th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,6,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 6
                                occupied += 1
    wb.save(filename)

def allocate_remaining_7thchoice(filename): #Allocate remaining students to their 7th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,7,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 7
                                occupied += 1
    wb.save(filename)

def allocate_remaining_8thchoice(filename): #Allocate remaining students to their 8th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,8,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 8
                                occupied += 1
    wb.save(filename)

def allocate_remaining_9thchoice(filename): #Allocate remaining students to their 7th choice CCA if CCAs still have vacancies
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_choice_dict = remaining_people(filename,9,"")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        occupied = int(value[1])
        quota = int(value[0])
        for cca in remaining_cca_lst:
            if cca == key and quota != occupied:
                for name, choice in remaining_choice_dict.items():
                    if choice[0] == cca:
                        for i in range(1,class_sheet.max_row):
                            if name == class_sheet.cell(row = i+1, column = 4).value and occupied != quota:
                                class_sheet["G" + str(i+1)] = cca
                                class_sheet["H" + str(i+1)] = 9
                                occupied += 1
    wb.save(filename)

def allocate_remaining(filename): #Randomly allocate students without CCAs to CCAs with available space
    wb = openpyxl.load_workbook(filename)
    class_sheet = wb["classlist"]
    remaining_name_lst = remaining_people(filename,0,"list")
    ccaquota_dict = extra_cca(filename,"quota")
    remaining_cca_lst = remaining_ccalist(ccaquota_dict)
    for key, value in ccaquota_dict.items():
        for cca in remaining_cca_lst:
            occupied = int(value[1])
            quota = int(value[0])
            if cca == key and occupied < quota:
                while occupied < quota:
                    name = random.choice(remaining_name_lst)
                    for i in range(1,class_sheet.max_row):
                        if name == class_sheet.cell(row = i+1, column = 4).value:
                            class_sheet["G" + str(i+1)] = cca
                            remaining_name_lst.remove(name)
                            occupied += 1
    for i in range(1,class_sheet.max_row): #Allocate the last student to Cricket as done in sample allocation
        if class_sheet.cell(row = i+1, column = 7).value == None:
            class_sheet["G" + str(i+1)] = "CRI"
    wb.save(filename)

allocate_full("unallocated2.xlsx")
